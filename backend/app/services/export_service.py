"""数据导出服务"""
import io
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select, func, distinct
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from ..models import Tool, User, Category, ClickLog, UserFavorite, UserLike, ToolFeedback
import logging

logger = logging.getLogger(__name__)


class ExportService:
    """数据导出服务"""

    def __init__(self, db: AsyncSession):
        self.db = db

    async def export_tools_stats(self, days: int = 30) -> bytes:
        """导出工具点击统计报表（与页面字段一致）"""
        now = datetime.now()
        current_start = now - timedelta(days=days)
        previous_start = current_start - timedelta(days=days)

        # 当前周期统计
        current_query = (
            select(
                Tool.id,
                Tool.name,
                Tool.provider,
                func.count(ClickLog.id).label("pv"),
                func.count(distinct(ClickLog.user_id)).label("uv"),
            )
            .join(ClickLog, Tool.id == ClickLog.tool_id)
            .where(ClickLog.clicked_at >= current_start)
            .group_by(Tool.id, Tool.name, Tool.provider)
            .order_by(func.count(ClickLog.id).desc())
        )
        current_result = await self.db.execute(current_query)
        current_rows = current_result.all()

        # 获取上周期数据用于计算环比
        tool_ids = [row.id for row in current_rows]
        previous_stats = {}
        if tool_ids:
            previous_query = (
                select(
                    ClickLog.tool_id,
                    func.count(ClickLog.id).label("pv"),
                    func.count(distinct(ClickLog.user_id)).label("uv"),
                )
                .where(
                    ClickLog.tool_id.in_(tool_ids),
                    ClickLog.clicked_at >= previous_start,
                    ClickLog.clicked_at < current_start
                )
                .group_by(ClickLog.tool_id)
            )
            previous_result = await self.db.execute(previous_query)
            for row in previous_result.all():
                previous_stats[row.tool_id] = {"pv": row.pv, "uv": row.uv}

        def calc_trend(current, previous):
            if previous == 0:
                return 100.0 if current > 0 else 0.0
            return round((current - previous) / previous * 100, 1)

        # 创建 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "工具点击统计"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 写入表头（与页面一致）
        headers = ["排名", "工具名称", "提供者", "PV", "UV", "PV环比(%)", "UV环比(%)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入数据
        for row_idx, row in enumerate(current_rows, 2):
            prev = previous_stats.get(row.id, {"pv": 0, "uv": 0})
            pv_trend = calc_trend(row.pv, prev["pv"])
            uv_trend = calc_trend(row.uv, prev["uv"])

            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
            ws.cell(row=row_idx, column=2, value=row.name).border = thin_border
            ws.cell(row=row_idx, column=3, value=row.provider or "-").border = thin_border
            ws.cell(row=row_idx, column=4, value=row.pv or 0).border = thin_border
            ws.cell(row=row_idx, column=5, value=row.uv or 0).border = thin_border
            ws.cell(row=row_idx, column=6, value=pv_trend).border = thin_border
            ws.cell(row=row_idx, column=7, value=uv_trend).border = thin_border

        # 调整列宽
        column_widths = [8, 25, 15, 12, 12, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    async def export_users_stats(self, days: int = 30) -> bytes:
        """导出用户分析报表（与页面字段一致）"""
        now = datetime.now()
        current_start = now - timedelta(days=days)
        previous_start = current_start - timedelta(days=days)

        # 当前周期统计
        current_query = (
            select(
                User.id,
                User.name,
                func.count(ClickLog.id).label("click_count"),
                func.max(ClickLog.clicked_at).label("last_click"),
            )
            .join(ClickLog, User.id == ClickLog.user_id)
            .where(ClickLog.clicked_at >= current_start)
            .group_by(User.id, User.name)
            .order_by(func.count(ClickLog.id).desc())
        )
        current_result = await self.db.execute(current_query)
        current_rows = current_result.all()

        # 获取上周期数据用于计算环比
        user_ids = [row.id for row in current_rows]
        previous_stats = {}
        if user_ids:
            previous_query = (
                select(
                    ClickLog.user_id,
                    func.count(ClickLog.id).label("click_count"),
                )
                .where(
                    ClickLog.user_id.in_(user_ids),
                    ClickLog.clicked_at >= previous_start,
                    ClickLog.clicked_at < current_start
                )
                .group_by(ClickLog.user_id)
            )
            previous_result = await self.db.execute(previous_query)
            for row in previous_result.all():
                previous_stats[row.user_id] = row.click_count

        def calc_trend(current, previous):
            if previous == 0:
                return 100.0 if current > 0 else 0.0
            return round((current - previous) / previous * 100, 1)

        def format_datetime(dt):
            if not dt:
                return ""
            if isinstance(dt, str):
                return dt[:16]
            return dt.strftime("%Y-%m-%d %H:%M")

        # 创建 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "用户分析"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 写入表头（与页面一致）
        headers = ["排名", "用户", "点击次数", "环比(%)", "最后访问"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入数据
        for row_idx, row in enumerate(current_rows, 2):
            prev_clicks = previous_stats.get(row.id, 0)
            click_trend = calc_trend(row.click_count, prev_clicks)

            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
            ws.cell(row=row_idx, column=2, value=row.name or "未知用户").border = thin_border
            ws.cell(row=row_idx, column=3, value=row.click_count or 0).border = thin_border
            ws.cell(row=row_idx, column=4, value=click_trend).border = thin_border
            ws.cell(row=row_idx, column=5, value=format_datetime(row.last_click)).border = thin_border

        # 调整列宽
        column_widths = [8, 20, 12, 12, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    async def export_trend_stats(self, days: int = 30) -> bytes:
        """导出趋势统计报表"""
        start_date = datetime.now() - timedelta(days=days)

        # 查询趋势数据
        query = (
            select(
                func.date(ClickLog.clicked_at).label("date"),
                func.count(ClickLog.id).label("pv"),
                func.count(distinct(ClickLog.user_id)).label("uv"),
            )
            .where(ClickLog.clicked_at >= start_date)
            .group_by(func.date(ClickLog.clicked_at))
            .order_by(func.date(ClickLog.clicked_at))
        )

        result = await self.db.execute(query)
        rows = result.all()

        # 创建 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "趋势统计"

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 写入表头
        headers = ["日期", "PV(浏览量)", "UV(独立访客)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入数据
        for row_idx, row in enumerate(rows, 2):
            date_str = row.date if isinstance(row.date, str) else row.date.isoformat()
            ws.cell(row=row_idx, column=1, value=date_str).border = thin_border
            ws.cell(row=row_idx, column=2, value=row.pv or 0).border = thin_border
            ws.cell(row=row_idx, column=3, value=row.uv or 0).border = thin_border

        # 调整列宽
        column_widths = [15, 15, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    async def export_interactions_stats(self) -> bytes:
        """导出工具互动统计报表"""
        # 收藏统计子查询
        fav_subq = (
            select(
                UserFavorite.tool_id,
                func.count(UserFavorite.id).label("fav_count")
            )
            .group_by(UserFavorite.tool_id)
            .subquery()
        )
        # 点赞统计子查询
        like_subq = (
            select(
                UserLike.tool_id,
                func.count(UserLike.id).label("like_count")
            )
            .group_by(UserLike.tool_id)
            .subquery()
        )

        query = (
            select(
                Tool.id,
                Tool.name,
                Tool.provider,
                func.coalesce(fav_subq.c.fav_count, 0).label("favorite_count"),
                func.coalesce(like_subq.c.like_count, 0).label("like_count"),
            )
            .outerjoin(fav_subq, Tool.id == fav_subq.c.tool_id)
            .outerjoin(like_subq, Tool.id == like_subq.c.tool_id)
            .where(Tool.is_active == True)
            .order_by(
                (func.coalesce(fav_subq.c.fav_count, 0) + func.coalesce(like_subq.c.like_count, 0)).desc()
            )
        )

        result = await self.db.execute(query)
        rows = result.all()

        # 创建 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "工具互动统计"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = ["工具ID", "工具名称", "提供者", "收藏数", "点赞数", "总计"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, row in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=row.id).border = thin_border
            ws.cell(row=row_idx, column=2, value=row.name).border = thin_border
            ws.cell(row=row_idx, column=3, value=row.provider or "-").border = thin_border
            ws.cell(row=row_idx, column=4, value=row.favorite_count).border = thin_border
            ws.cell(row=row_idx, column=5, value=row.like_count).border = thin_border
            ws.cell(row=row_idx, column=6, value=row.favorite_count + row.like_count).border = thin_border

        column_widths = [10, 25, 15, 12, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    async def export_providers_stats(self) -> bytes:
        """导出提供者统计报表"""
        query = (
            select(
                Tool.provider,
                func.count(Tool.id).label("tool_count"),
                func.count(ClickLog.id).label("click_count"),
            )
            .outerjoin(ClickLog, Tool.id == ClickLog.tool_id)
            .where(Tool.is_active == True)
            .where(Tool.provider.isnot(None))
            .where(Tool.provider != "")
            .group_by(Tool.provider)
            .order_by(func.count(Tool.id).desc())
        )

        result = await self.db.execute(query)
        rows = result.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "提供者统计"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = ["提供者", "贡献工具数", "总点击数", "平均点击"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, row in enumerate(rows, 2):
            avg_clicks = round(row.click_count / row.tool_count) if row.tool_count > 0 else 0
            ws.cell(row=row_idx, column=1, value=row.provider).border = thin_border
            ws.cell(row=row_idx, column=2, value=row.tool_count).border = thin_border
            ws.cell(row=row_idx, column=3, value=row.click_count).border = thin_border
            ws.cell(row=row_idx, column=4, value=avg_clicks).border = thin_border

        column_widths = [20, 15, 15, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    async def export_wants_stats(self) -> bytes:
        """导出用户想要统计报表"""
        query = (
            select(
                ToolFeedback.tool_name,
                func.count(ToolFeedback.id).label("want_count"),
            )
            .where(ToolFeedback.feedback_type == "want")
            .where(ToolFeedback.tool_name.isnot(None))
            .where(ToolFeedback.tool_name != "")
            .group_by(ToolFeedback.tool_name)
            .order_by(func.count(ToolFeedback.id).desc())
        )

        result = await self.db.execute(query)
        rows = result.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "用户想要"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = ["工具名称", "想要次数"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, row in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=row.tool_name).border = thin_border
            ws.cell(row=row_idx, column=2, value=row.want_count).border = thin_border

        column_widths = [30, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
