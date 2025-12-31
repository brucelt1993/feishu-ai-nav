"""Excel导入服务"""
from io import BytesIO
from typing import Any
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
import logging

from ..models import Tool, Category

logger = logging.getLogger(__name__)


class ImportResult:
    """导入结果"""
    def __init__(self):
        self.created = 0
        self.updated = 0
        self.skipped = 0
        self.errors: list[str] = []

    def to_dict(self):
        return {
            "created": self.created,
            "updated": self.updated,
            "skipped": self.skipped,
            "errors": self.errors,
            "total": self.created + self.updated + self.skipped,
        }


# Excel列映射
COLUMN_MAPPING = {
    "名称": "name",
    "描述": "description",
    "链接": "target_url",
    "图标URL": "icon_url",
    "提供者": "provider",  # 谁推荐了这个工具
    "分类": "category_name",  # 通过名称匹配分类ID
    "排序": "sort_order",
}

REQUIRED_COLUMNS = ["名称", "链接"]


async def parse_excel(file_content: bytes) -> list[dict[str, Any]]:
    """解析Excel文件"""
    wb = load_workbook(BytesIO(file_content), read_only=True)
    ws = wb.active

    if not ws:
        raise ValueError("Excel文件为空")

    # 读取表头
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        raise ValueError("Excel文件为空")

    headers = [str(h).strip() if h else "" for h in rows[0]]

    # 验证必要列
    for col in REQUIRED_COLUMNS:
        if col not in headers:
            raise ValueError(f"缺少必要列: {col}")

    # 解析数据行
    data = []
    for row_idx, row in enumerate(rows[1:], start=2):
        if not any(row):  # 跳过空行
            continue

        item = {"_row": row_idx}
        for col_idx, header in enumerate(headers):
            if header in COLUMN_MAPPING:
                value = row[col_idx] if col_idx < len(row) else None
                if value is not None:
                    value = str(value).strip() if value else ""
                item[COLUMN_MAPPING[header]] = value

        # 验证必填字段
        if not item.get("name"):
            continue
        if not item.get("target_url"):
            continue

        data.append(item)

    wb.close()
    return data


async def import_tools(
    db: AsyncSession,
    file_content: bytes,
    update_existing: bool = True,
) -> ImportResult:
    """
    导入工具数据

    Args:
        db: 数据库会话
        file_content: Excel文件内容
        update_existing: 是否更新已存在的工具（按名称匹配）

    Returns:
        ImportResult: 导入结果
    """
    result = ImportResult()

    try:
        data = await parse_excel(file_content)
    except Exception as e:
        result.errors.append(f"解析Excel失败: {str(e)}")
        return result

    if not data:
        result.errors.append("没有有效数据")
        return result

    # 获取所有分类（用于名称匹配）
    categories_result = await db.execute(select(Category))
    categories = {c.name: c.id for c in categories_result.scalars().all()}

    for item in data:
        row_num = item.pop("_row", "?")
        name = item.get("name", "").strip()

        if not name:
            result.skipped += 1
            continue

        try:
            # 处理分类
            category_name = item.pop("category_name", None)
            if category_name and category_name in categories:
                item["category_id"] = categories[category_name]

            # 处理排序
            if "sort_order" in item:
                try:
                    item["sort_order"] = int(item["sort_order"])
                except (ValueError, TypeError):
                    item["sort_order"] = 0

            # 查找是否已存在
            existing = await db.execute(
                select(Tool).where(Tool.name == name)
            )
            tool = existing.scalar_one_or_none()

            if tool:
                if update_existing:
                    # 更新已有工具
                    for key, value in item.items():
                        if value is not None and value != "":
                            setattr(tool, key, value)
                    result.updated += 1
                    logger.info(f"更新工具: {name}")
                else:
                    result.skipped += 1
            else:
                # 创建新工具
                tool = Tool(
                    name=name,
                    description=item.get("description", ""),
                    target_url=item.get("target_url", ""),
                    icon_url=item.get("icon_url", ""),
                    provider=item.get("provider", ""),
                    category_id=item.get("category_id"),
                    sort_order=item.get("sort_order", 0),
                    is_active=True,
                )
                db.add(tool)
                result.created += 1
                logger.info(f"创建工具: {name}")

        except Exception as e:
            error_msg = f"第{row_num}行 [{name}] 处理失败: {str(e)}"
            result.errors.append(error_msg)
            logger.error(error_msg)

    await db.commit()
    return result


def generate_template() -> bytes:
    """生成导入模板"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "工具导入"

    # 写入表头
    headers = ["名称", "描述", "链接", "图标URL", "提供者", "分类", "排序"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # 写入示例数据
    example = ["ChatGPT", "OpenAI的对话AI", "https://chat.openai.com", "", "张三", "AI助手", "1"]
    for col, value in enumerate(example, start=1):
        ws.cell(row=2, column=col, value=value)

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
