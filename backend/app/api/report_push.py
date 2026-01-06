"""报表推送管理API"""
import io
import logging
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from pydantic import BaseModel, EmailStr

from app.database import get_db
from app.models import ReportPushSettings, ReportRecipient, ReportPushHistory
from app.api.admin import verify_admin
from app.services.stats_service import StatsService
from app.services.feishu_service import feishu_service

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/admin/report-push", tags=["报表推送"])


# ============ Schemas ============
class SettingsUpdate(BaseModel):
    enabled: bool = False
    push_time: Optional[str] = None
    report_types: Optional[List[str]] = None
    days: Optional[int] = 7


class RecipientCreate(BaseModel):
    name: str
    email: EmailStr


class RecipientUpdate(BaseModel):
    is_active: Optional[bool] = None


class PushRequest(BaseModel):
    report_types: List[str]  # clicks, interactions, providers, users, wants, custom
    days: int = 7
    method: str = "feishu"  # feishu or email
    chat_ids: Optional[List[str]] = None  # 群聊ID列表
    custom_content: Optional[str] = None  # 自定义报表内容


class PreviewRequest(BaseModel):
    report_types: List[str]  # clicks, interactions, providers, users, wants, custom
    days: int = 7
    custom_content: Optional[str] = None  # 自定义报表内容


# ============ 设置管理 ============
@router.get("/settings")
async def get_settings(
    db: AsyncSession = Depends(get_db),
    _admin: str = Depends(verify_admin)
):
    """获取推送设置"""
    result = await db.execute(select(ReportPushSettings).limit(1))
    settings = result.scalar_one_or_none()

    if not settings:
        return {
            "enabled": False,
            "push_time": None,
            "report_types": ["overview", "tools"],
            "days": 7
        }

    return {
        "enabled": settings.enabled,
        "push_time": settings.push_time,
        "report_types": settings.report_types.split(",") if settings.report_types else ["overview", "tools"],
        "days": settings.days or 7
    }


@router.put("/settings")
async def update_settings(
    data: SettingsUpdate,
    db: AsyncSession = Depends(get_db),
    _admin: str = Depends(verify_admin)
):
    """更新推送设置"""
    result = await db.execute(select(ReportPushSettings).limit(1))
    settings = result.scalar_one_or_none()

    report_types_str = ",".join(data.report_types) if data.report_types else "overview,tools"

    if not settings:
        settings = ReportPushSettings(
            enabled=data.enabled,
            push_time=data.push_time,
            report_types=report_types_str,
            days=data.days or 7
        )
        db.add(settings)
    else:
        settings.enabled = data.enabled
        settings.push_time = data.push_time
        settings.report_types = report_types_str
        settings.days = data.days or 7
        settings.updated_at = datetime.utcnow()

    await db.commit()
    return {"message": "设置已更新"}


# ============ 接收人管理 ============
@router.get("/recipients")
async def get_recipients(
    db: AsyncSession = Depends(get_db),
    _admin: str = Depends(verify_admin)
):
    """获取推送接收人列表"""
    result = await db.execute(
        select(ReportRecipient).order_by(ReportRecipient.created_at.desc())
    )
    recipients = result.scalars().all()

    return [
        {
            "id": r.id,
            "name": r.name,
            "email": r.email,
            "is_active": r.is_active,
            "created_at": r.created_at.isoformat() if r.created_at else None
        }
        for r in recipients
    ]


@router.post("/recipients")
async def add_recipient(
    data: RecipientCreate,
    db: AsyncSession = Depends(get_db),
    _admin: str = Depends(verify_admin)
):
    """添加推送接收人"""
    # 检查邮箱是否已存在
    result = await db.execute(
        select(ReportRecipient).where(ReportRecipient.email == data.email)
    )
    if result.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="该邮箱已存在")

    recipient = ReportRecipient(name=data.name, email=data.email)
    db.add(recipient)
    await db.commit()
    await db.refresh(recipient)

    return {
        "id": recipient.id,
        "name": recipient.name,
        "email": recipient.email,
        "is_active": recipient.is_active,
        "created_at": recipient.created_at.isoformat()
    }


@router.put("/recipients/{recipient_id}")
async def update_recipient(
    recipient_id: int,
    data: RecipientUpdate,
    db: AsyncSession = Depends(get_db),
    _admin: str = Depends(verify_admin)
):
    """更新推送接收人"""
    result = await db.execute(
        select(ReportRecipient).where(ReportRecipient.id == recipient_id)
    )
    recipient = result.scalar_one_or_none()

    if not recipient:
        raise HTTPException(status_code=404, detail="接收人不存在")

    if data.is_active is not None:
        recipient.is_active = data.is_active

    recipient.updated_at = datetime.utcnow()
    await db.commit()

    return {"message": "更新成功"}


@router.delete("/recipients/{recipient_id}")
async def delete_recipient(
    recipient_id: int,
    db: AsyncSession = Depends(get_db),
    _admin: str = Depends(verify_admin)
):
    """删除推送接收人"""
    result = await db.execute(
        select(ReportRecipient).where(ReportRecipient.id == recipient_id)
    )
    recipient = result.scalar_one_or_none()

    if not recipient:
        raise HTTPException(status_code=404, detail="接收人不存在")

    await db.delete(recipient)
    await db.commit()

    return {"message": "删除成功"}


# ============ 群聊管理 ============
@router.get("/chats")
async def get_bot_chats(
    _admin: str = Depends(verify_admin)
):
    """获取机器人已加入的群聊列表"""
    try:
        chats = await feishu_service.get_bot_joined_chats()
        return chats
    except Exception as e:
        logger.error(f"获取群聊列表失败: {e}")
        raise HTTPException(status_code=500, detail=f"获取群聊列表失败: {str(e)}")


# ============ 推送历史 ============
@router.get("/history")
async def get_history(
    page: int = 1,
    size: int = 10,
    db: AsyncSession = Depends(get_db),
    _admin: str = Depends(verify_admin)
):
    """获取推送历史"""
    offset = (page - 1) * size

    # 总数
    count_result = await db.execute(select(func.count(ReportPushHistory.id)))
    total = count_result.scalar()

    # 列表
    result = await db.execute(
        select(ReportPushHistory)
        .order_by(ReportPushHistory.pushed_at.desc())
        .offset(offset)
        .limit(size)
    )
    items = result.scalars().all()

    return {
        "total": total,
        "items": [
            {
                "id": h.id,
                "report_type": h.report_type,
                "push_method": h.push_method,
                "recipient_count": h.recipient_count,
                "status": h.status,
                "error_msg": h.error_msg,
                "pushed_at": h.pushed_at.isoformat() if h.pushed_at else None
            }
            for h in items
        ]
    }


# ============ 推送和预览 ============
@router.post("/preview")
async def preview_report(
    data: PreviewRequest,
    db: AsyncSession = Depends(get_db),
    _admin: str = Depends(verify_admin)
):
    """预览报表数据"""
    stats_service = StatsService(db)
    result = {}

    # 工具点击
    if "clicks" in data.report_types:
        result["clicks"] = await stats_service.get_tool_stats(days=data.days, limit=10)

    # 工具互动
    if "interactions" in data.report_types:
        result["interactions"] = await stats_service.get_tool_interactions(limit=10)

    # 提供者排行
    if "providers" in data.report_types:
        result["providers"] = await stats_service.get_provider_stats(limit=10)

    # 用户分析
    if "users" in data.report_types:
        result["users"] = await stats_service.get_user_stats(days=data.days, limit=10)

    # 用户想要
    if "wants" in data.report_types:
        result["wants"] = await stats_service.get_want_list(limit=10)

    # 自定义报表
    if "custom" in data.report_types and data.custom_content:
        result["custom"] = {"content": data.custom_content}

    return result


@router.post("/push")
async def push_report(
    data: PushRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    _admin: str = Depends(verify_admin)
):
    """推送报表"""
    # 获取活跃的接收人
    result = await db.execute(
        select(ReportRecipient).where(ReportRecipient.is_active == True)
    )
    recipient_objs = result.scalars().all()

    # 转换为字典
    recipients = [{"name": r.name, "email": r.email, "type": "user"} for r in recipient_objs]

    # 添加群聊接收人
    chat_ids = data.chat_ids or []
    for chat_id in chat_ids:
        recipients.append({"chat_id": chat_id, "type": "chat"})

    if not recipients:
        raise HTTPException(status_code=400, detail="没有可用的推送接收人或群聊")

    # 获取报表数据
    stats_service = StatsService(db)
    report_data = {}

    # 工具点击
    if "clicks" in data.report_types:
        report_data["clicks"] = await stats_service.get_tool_stats(days=data.days, limit=10)

    # 工具互动
    if "interactions" in data.report_types:
        report_data["interactions"] = await stats_service.get_tool_interactions(limit=10)

    # 提供者排行
    if "providers" in data.report_types:
        report_data["providers"] = await stats_service.get_provider_stats(limit=10)

    # 用户分析
    if "users" in data.report_types:
        report_data["users"] = await stats_service.get_user_stats(days=data.days, limit=10)

    # 用户想要
    if "wants" in data.report_types:
        report_data["wants"] = await stats_service.get_want_list(limit=10)

    # 自定义报表
    if "custom" in data.report_types and data.custom_content:
        report_data["custom"] = {"content": data.custom_content}

    # 记录推送历史
    history = ReportPushHistory(
        report_type=",".join(data.report_types),
        push_method=data.method,
        recipient_count=len(recipients),
        status="pending"
    )
    db.add(history)
    await db.commit()
    await db.refresh(history)

    # 后台推送
    if data.method == "feishu":
        background_tasks.add_task(
            push_feishu_report,
            history.id,
            recipients,
            report_data,
            data.days
        )
    else:
        background_tasks.add_task(
            push_email_report,
            history.id,
            recipients,
            report_data,
            data.days
        )

    return {"message": "推送任务已提交", "history_id": history.id}


async def push_feishu_report(history_id: int, recipients: list, report_data: dict, days: int):
    """飞书消息推送"""
    from app.database import async_session

    async with async_session() as db:
        try:
            # 构建卡片消息
            card = build_report_card(report_data, days)

            # 逐个推送给接收人
            success_count = 0
            errors = []

            for recipient in recipients:
                try:
                    if recipient.get("type") == "chat":
                        # 推送到群聊
                        await feishu_service.send_card_message(
                            recipient["chat_id"],
                            card,
                            receive_id_type="chat_id"
                        )
                        success_count += 1
                    else:
                        # 推送到个人（通过邮箱获取open_id）
                        user_info = await feishu_service.get_user_by_email(recipient["email"])
                        if user_info and user_info.get("open_id"):
                            await feishu_service.send_card_message(
                                user_info["open_id"],
                                card,
                                receive_id_type="open_id"
                            )
                            success_count += 1
                        else:
                            errors.append(f"{recipient['name']}: 未找到飞书用户")
                except Exception as e:
                    name = recipient.get("name") or recipient.get("chat_id", "未知")
                    errors.append(f"{name}: {str(e)}")
                    logger.error(f"推送给 {name} 失败: {e}")

            # 更新历史记录
            result = await db.execute(
                select(ReportPushHistory).where(ReportPushHistory.id == history_id)
            )
            history = result.scalar_one_or_none()
            if history:
                if success_count > 0:
                    history.status = "success"
                    if errors:
                        history.error_msg = f"部分成功({success_count}/{len(recipients)}): " + "; ".join(errors)
                else:
                    history.status = "failed"
                    history.error_msg = "; ".join(errors)
                await db.commit()

            logger.info(f"飞书推送完成: {success_count}/{len(recipients)}")

        except Exception as e:
            logger.error(f"飞书推送失败: {e}")
            result = await db.execute(
                select(ReportPushHistory).where(ReportPushHistory.id == history_id)
            )
            history = result.scalar_one_or_none()
            if history:
                history.status = "failed"
                history.error_msg = str(e)
                await db.commit()


async def push_email_report(history_id: int, recipients: list, report_data: dict, days: int):
    """邮件推送(Excel附件)"""
    from app.database import async_session

    # 过滤只保留用户类型的接收人（群聊不支持邮件推送）
    user_recipients = [r for r in recipients if r.get("type") == "user"]
    if not user_recipients:
        logger.warning("没有可用的邮件接收人（群聊不支持邮件推送）")
        return

    async with async_session() as db:
        try:
            # 生成Excel文件
            excel_content = generate_report_excel(report_data, days)

            # 逐个发送邮件
            success_count = 0
            errors = []

            for recipient in user_recipients:
                try:
                    # 通过飞书邮件API发送
                    await feishu_service.send_email_with_attachment(
                        to_email=recipient["email"],
                        subject=f"AI工具导航统计报表 - 近{days}天",
                        content="请查看附件中的统计报表。",
                        attachment_name=f"report_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        attachment_content=excel_content
                    )
                    success_count += 1
                except Exception as e:
                    errors.append(f"{recipient['name']}: {str(e)}")
                    logger.error(f"发送邮件给 {recipient['name']} 失败: {e}")

            # 更新历史记录
            result = await db.execute(
                select(ReportPushHistory).where(ReportPushHistory.id == history_id)
            )
            history = result.scalar_one_or_none()
            if history:
                if success_count > 0:
                    history.status = "success"
                    if errors:
                        history.error_msg = f"部分成功({success_count}/{len(recipients)}): " + "; ".join(errors)
                else:
                    history.status = "failed"
                    history.error_msg = "; ".join(errors)
                await db.commit()

            logger.info(f"邮件推送完成: {success_count}/{len(recipients)}")

        except Exception as e:
            logger.error(f"邮件推送失败: {e}")
            result = await db.execute(
                select(ReportPushHistory).where(ReportPushHistory.id == history_id)
            )
            history = result.scalar_one_or_none()
            if history:
                history.status = "failed"
                history.error_msg = str(e)
                await db.commit()


def build_report_card(report_data: dict, days: int) -> dict:
    """构建飞书卡片消息"""
    elements = []
    is_custom_only = "custom" in report_data and len(report_data) == 1

    # 自定义内容单独推送时，不显示统计报表标题
    if is_custom_only:
        custom = report_data["custom"]
        elements.append({
            "tag": "markdown",
            "content": custom.get('content', '')
        })
        return {
            "config": {"wide_screen_mode": True},
            "header": {
                "title": {"content": "📢 通知", "tag": "plain_text"},
                "template": "orange"
            },
            "elements": elements
        }

    # 有统计数据时，显示报表标题
    elements.append({
        "tag": "markdown",
        "content": f"**📊 AI工具导航统计报表（近{days}天）**"
    })

    # 自定义内容放在报表数据前面
    if "custom" in report_data:
        custom = report_data["custom"]
        elements.append({"tag": "hr"})
        elements.append({
            "tag": "markdown",
            "content": f"**📢 通知**\n{custom.get('content', '')}"
        })

    # 工具点击排行 - 表格形式
    if "clicks" in report_data and report_data["clicks"]:
        tools = report_data["clicks"][:5]
        elements.append({"tag": "hr"})
        elements.append({
            "tag": "markdown",
            "content": "**🔥 工具点击 TOP5**"
        })
        # 使用表格组件
        elements.append({
            "tag": "table",
            "page_size": 5,
            "row_height": "low",
            "header_style": {
                "text_align": "center",
                "text_size": "normal",
                "background_style": "grey",
                "text_color": "grey",
                "bold": True
            },
            "columns": [
                {"name": "rank", "display_name": "#", "width": "auto", "data_type": "text"},
                {"name": "tool", "display_name": "工具", "width": "auto", "data_type": "text"},
                {"name": "pv", "display_name": "PV", "width": "auto", "data_type": "number"},
                {"name": "uv", "display_name": "UV", "width": "auto", "data_type": "number"}
            ],
            "rows": [
                {"rank": str(i+1), "tool": t['tool_name'], "pv": t['click_count'], "uv": t['unique_users']}
                for i, t in enumerate(tools)
            ]
        })

    # 工具互动排行 - 表格形式
    if "interactions" in report_data and report_data["interactions"]:
        items = report_data["interactions"][:5]
        elements.append({"tag": "hr"})
        elements.append({
            "tag": "markdown",
            "content": "**💫 工具互动 TOP5**"
        })
        elements.append({
            "tag": "table",
            "page_size": 5,
            "row_height": "low",
            "header_style": {
                "text_align": "center",
                "text_size": "normal",
                "background_style": "grey",
                "text_color": "grey",
                "bold": True
            },
            "columns": [
                {"name": "rank", "display_name": "#", "width": "auto", "data_type": "text"},
                {"name": "tool", "display_name": "工具", "width": "auto", "data_type": "text"},
                {"name": "fav", "display_name": "⭐收藏", "width": "auto", "data_type": "number"},
                {"name": "like", "display_name": "👍点赞", "width": "auto", "data_type": "number"}
            ],
            "rows": [
                {"rank": str(i+1), "tool": t['tool_name'], "fav": t['favorite_count'], "like": t['like_count']}
                for i, t in enumerate(items)
            ]
        })

    # 提供者排行 - 表格形式
    if "providers" in report_data and report_data["providers"]:
        items = report_data["providers"][:5]
        elements.append({"tag": "hr"})
        elements.append({
            "tag": "markdown",
            "content": "**🏆 提供者排行 TOP5**"
        })
        elements.append({
            "tag": "table",
            "page_size": 5,
            "row_height": "low",
            "header_style": {
                "text_align": "center",
                "text_size": "normal",
                "background_style": "grey",
                "text_color": "grey",
                "bold": True
            },
            "columns": [
                {"name": "rank", "display_name": "#", "width": "auto", "data_type": "text"},
                {"name": "provider", "display_name": "提供者", "width": "auto", "data_type": "text"},
                {"name": "tools", "display_name": "工具数", "width": "auto", "data_type": "number"},
                {"name": "clicks", "display_name": "点击数", "width": "auto", "data_type": "number"}
            ],
            "rows": [
                {"rank": str(i+1), "provider": t['provider'], "tools": t['tool_count'], "clicks": t['click_count']}
                for i, t in enumerate(items)
            ]
        })

    # 用户排行 - 表格形式
    if "users" in report_data and report_data["users"]:
        users = report_data["users"][:5]
        elements.append({"tag": "hr"})
        elements.append({
            "tag": "markdown",
            "content": "**👥 活跃用户 TOP5**"
        })
        elements.append({
            "tag": "table",
            "page_size": 5,
            "row_height": "low",
            "header_style": {
                "text_align": "center",
                "text_size": "normal",
                "background_style": "grey",
                "text_color": "grey",
                "bold": True
            },
            "columns": [
                {"name": "rank", "display_name": "#", "width": "auto", "data_type": "text"},
                {"name": "user", "display_name": "用户", "width": "auto", "data_type": "text"},
                {"name": "clicks", "display_name": "点击次数", "width": "auto", "data_type": "number"}
            ],
            "rows": [
                {"rank": str(i+1), "user": u['user_name'], "clicks": u['click_count']}
                for i, u in enumerate(users)
            ]
        })

    # 用户想要 - 表格形式
    if "wants" in report_data and report_data["wants"]:
        items = report_data["wants"][:5]
        elements.append({"tag": "hr"})
        elements.append({
            "tag": "markdown",
            "content": "**💡 用户想要 TOP5**"
        })
        elements.append({
            "tag": "table",
            "page_size": 5,
            "row_height": "low",
            "header_style": {
                "text_align": "center",
                "text_size": "normal",
                "background_style": "grey",
                "text_color": "grey",
                "bold": True
            },
            "columns": [
                {"name": "rank", "display_name": "#", "width": "auto", "data_type": "text"},
                {"name": "tool", "display_name": "工具", "width": "auto", "data_type": "text"},
                {"name": "count", "display_name": "想要人数", "width": "auto", "data_type": "number"}
            ],
            "rows": [
                {"rank": str(i+1), "tool": t['tool_name'], "count": t['want_count']}
                for i, t in enumerate(items)
            ]
        })

    return {
        "config": {"wide_screen_mode": True},
        "header": {
            "title": {"content": "📊 统计报表", "tag": "plain_text"},
            "template": "blue"
        },
        "elements": elements
    }


def generate_report_excel(report_data: dict, days: int) -> bytes:
    """生成Excel报表"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel库未安装")

    wb = openpyxl.Workbook()
    first_sheet = True

    def style_headers(ws, headers):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")

    # 工具点击Sheet
    if "clicks" in report_data and report_data["clicks"]:
        if first_sheet:
            ws = wb.active
            ws.title = "工具点击"
            first_sheet = False
        else:
            ws = wb.create_sheet("工具点击")
        headers = ["排名", "工具名称", "提供者", "PV", "UV", "PV环比(%)", "UV环比(%)"]
        style_headers(ws, headers)
        for row, tool in enumerate(report_data["clicks"], 2):
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2, value=tool.get("tool_name", ""))
            ws.cell(row=row, column=3, value=tool.get("provider", "-"))
            ws.cell(row=row, column=4, value=tool.get("click_count", 0))
            ws.cell(row=row, column=5, value=tool.get("unique_users", 0))
            ws.cell(row=row, column=6, value=tool.get("pv_trend", 0))
            ws.cell(row=row, column=7, value=tool.get("uv_trend", 0))

    # 工具互动Sheet
    if "interactions" in report_data and report_data["interactions"]:
        if first_sheet:
            ws = wb.active
            ws.title = "工具互动"
            first_sheet = False
        else:
            ws = wb.create_sheet("工具互动")
        headers = ["排名", "工具名称", "提供者", "收藏数", "点赞数", "总计"]
        style_headers(ws, headers)
        for row, item in enumerate(report_data["interactions"], 2):
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2, value=item.get("tool_name", ""))
            ws.cell(row=row, column=3, value=item.get("provider", "-"))
            ws.cell(row=row, column=4, value=item.get("favorite_count", 0))
            ws.cell(row=row, column=5, value=item.get("like_count", 0))
            ws.cell(row=row, column=6, value=item.get("total", 0))

    # 提供者统计Sheet
    if "providers" in report_data and report_data["providers"]:
        if first_sheet:
            ws = wb.active
            ws.title = "提供者排行"
            first_sheet = False
        else:
            ws = wb.create_sheet("提供者排行")
        headers = ["排名", "提供者", "工具数", "点击数", "平均点击"]
        style_headers(ws, headers)
        for row, item in enumerate(report_data["providers"], 2):
            avg = round(item.get("click_count", 0) / item.get("tool_count", 1)) if item.get("tool_count", 0) > 0 else 0
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2, value=item.get("provider", ""))
            ws.cell(row=row, column=3, value=item.get("tool_count", 0))
            ws.cell(row=row, column=4, value=item.get("click_count", 0))
            ws.cell(row=row, column=5, value=avg)

    # 用户统计Sheet
    if "users" in report_data and report_data["users"]:
        if first_sheet:
            ws = wb.active
            ws.title = "用户分析"
            first_sheet = False
        else:
            ws = wb.create_sheet("用户分析")
        headers = ["排名", "用户", "点击次数", "环比(%)", "最后访问"]
        style_headers(ws, headers)
        for row, user in enumerate(report_data["users"], 2):
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2, value=user.get("user_name", ""))
            ws.cell(row=row, column=3, value=user.get("click_count", 0))
            ws.cell(row=row, column=4, value=user.get("click_trend", 0))
            ws.cell(row=row, column=5, value=user.get("last_click", ""))

    # 用户想要Sheet
    if "wants" in report_data and report_data["wants"]:
        if first_sheet:
            ws = wb.active
            ws.title = "用户想要"
            first_sheet = False
        else:
            ws = wb.create_sheet("用户想要")
        headers = ["排名", "工具名称", "想要次数"]
        style_headers(ws, headers)
        for row, item in enumerate(report_data["wants"], 2):
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2, value=item.get("tool_name", ""))
            ws.cell(row=row, column=3, value=item.get("want_count", 0))

    # 自定义内容Sheet
    if "custom" in report_data:
        if first_sheet:
            ws = wb.active
            ws.title = "自定义通知"
            first_sheet = False
        else:
            ws = wb.create_sheet("自定义通知")
        ws.cell(row=1, column=1, value="通知内容")
        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.cell(row=2, column=1, value=report_data["custom"].get("content", ""))

    # 保存到字节流
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
